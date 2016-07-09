from openpyxl import*
from flask import*
import sqlite3
import os
import datetime


def shutdown_server():
    func = request.environ.get('werkzeug.server.shutdown')
    if func is None:
        raise RuntimeError('Not running with the Werkzeug Server')
    func()



def serv1(das):
        app = Flask(__name__)

        @app.route('/')
        def index():
            return '''<!DOCTYPE html>
<html>
  <head>
    <title>Login page</title>
  </head>

  <body>

    <form action="/login" method="post">
       name: &nbsp&nbsp&nbsp&nbsp<input type="text" name="roll"></input><br><br>
       pass: &nbsp&nbsp&nbsp&nbsp&nbsp<input type="password" name="pass"></input><br><br>
       <input type="submit" value="Submit"></input>
    </form>

  </body>
</html>'''

        @app.route('/login', methods = ['POST'])
        def login():
            conn = sqlite3.connect('./DataBase/attendance.db')
            num = request.form['roll']
            pwd = request.form['pass']
            cur = conn.cursor()
            cur.execute("SELECT * FROM example WHERE roll=?", (num,))
            row = cur.fetchone()
            if(row):
                # (num1,pwd1) = row
                if(pwd == '1234'):
                    # cur.execute("ALTER TABLE example ADD COLUMN ?  DEFAULT 0")
                    now = datetime.datetime.now()
                    date = str(now.date())
                    cur.execute("UPDATE example SET "+'['+date+']'+"=(1) WHERE roll=?",(num,) )
                    conn.commit()
                    return "SUCCESS"
                else:
                    return "CHECK PASSWORD"
            else:
                return "CHECK ROLL NUMBER"
            # if int(request.form['pass'])==1234 :
            #     return form['roll']
            #     PATH='attendence.xlsx'
            #     if os.path.isfile(PATH):
            #         wb = load_workbook('attendence.xlsx')
            #     else:
            #         wb=Workbook()
            #     ws = wb.active
            #     today_column=ws.max_column
            #     column1=ws.columns[0]
            #     for j in range(1,len(column1)):
            #         exist=False
            #         if column1[j].value == request.form['roll']:
            #             value=j
            #             exist=True
            #             break
            #
            #
            #     if exist:
            #         ws.cell(row=ws.cell(column1[value].coordinate).row,column=today_column).value=1
            #         wb.save("attendence.xlsx")
            #         return 'SUCCESS'
            #     else:
            #         return 'CHECK YOUR ROLL NO.'
            # else:
            #      return 'WRONG PASSWORD'

        @app.route('/shutdown')
        def shutdown():
            shutdown_server()
            return 'Server shutting down...'


        app.run(host='0.0.0.0',port='8080',threaded=True)

serv1('s')
